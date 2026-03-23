"""
FAQ 知識庫管理 API
"""
import asyncio
import json
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.api.v1.auth import get_current_user, oauth2_scheme
from app.core.security import decode_access_token
from app.models.user import User
from app.services.faq_service import FaqService
from app.services.chatbot_service import chatbot_service as ai_chatbot
from app.schemas.faq import (
    FaqCategoryToggleSchema,
    FaqRuleCreateSchema,
    FaqRuleUpdateSchema,
    FaqRuleToggleSchema,
    AiTokenUsageUpdateSchema,
    FaqModuleAuthUpdateSchema,
    AiTestChatRequestSchema,
)
from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional
import csv
import io
import logging
import time
import openpyxl
import xlrd
import xlwt

router = APIRouter()
logger = logging.getLogger(__name__)
faq_service = FaqService()


# === 大分類 ===


@router.get("/categories", response_model=dict)
async def get_categories(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """取得大分類清單（含欄位定義與規則數量）"""
    industry = await faq_service.get_default_industry(db)
    if not industry:
        raise HTTPException(status_code=404, detail="尚未設定產業資料")

    categories = await faq_service.get_categories(db, industry.id)

    # Build response with PMS connection status for each category
    result = []
    for cat in categories:
        cat_data = {
            "id": cat.id,
            "industry_id": cat.industry_id,
            "name": cat.name,
            "is_active": cat.is_active,
            "data_source_type": cat.data_source_type,
            "is_system_default": cat.is_system_default,
            "sort_order": cat.sort_order,
            "fields": [
                {
                    "id": f.id,
                    "field_name": f.field_name,
                    "field_type": f.field_type,
                    "is_required": f.is_required,
                    "sort_order": f.sort_order,
                }
                for f in sorted(cat.fields, key=lambda x: x.sort_order)
            ],
            "rule_count": getattr(cat, "rule_count", 0),
            "published_count": getattr(cat, "published_count", 0),
            "updated_at": cat.updated_at.isoformat() if cat.updated_at else None,
            "pms_connection": None,
        }
        # Attach PMS connection info if data_source_type is pms
        if cat.data_source_type == "pms":
            pms_conn = await faq_service.get_pms_connection(db, cat.id)
            if pms_conn:
                cat_data["pms_connection"] = {
                    "status": pms_conn.status,
                    "last_synced_at": pms_conn.last_synced_at.isoformat() if pms_conn.last_synced_at else None,
                }
        result.append(cat_data)

    return {
        "code": 200,
        "message": "查詢成功",
        "data": result,
    }


@router.patch("/categories/{category_id}/toggle", response_model=dict)
async def toggle_category(
    category_id: int,
    data: FaqCategoryToggleSchema,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """啟用/停用大分類"""
    category = await faq_service.toggle_category(db, category_id, data.is_active)
    if not category:
        raise HTTPException(status_code=404, detail="大分類不存在")

    _bump_rule_modified()
    return {
        "code": 200,
        "message": f"大分類已{'啟用' if data.is_active else '停用'}",
        "data": {"id": category.id, "is_active": category.is_active},
    }


# === 規則 CRUD ===


@router.get("/categories/{category_id}/rules", response_model=dict)
async def get_rules(
    category_id: int,
    status: Optional[str] = Query(None, description="狀態篩選：draft/active/disabled"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """取得分類下規則清單"""
    result = await faq_service.get_rules(db, category_id, status, page, page_size)

    items = []
    for rule in result["items"]:
        content = rule.content_json
        if isinstance(content, str):
            try:
                content = json.loads(content)
            except (json.JSONDecodeError, TypeError):
                pass
        items.append({
            "id": rule.id,
            "category_id": rule.category_id,
            "content_json": content,
            "status": rule.status,
            "is_enabled": rule.is_enabled,
            "published_at": rule.published_at.isoformat() if rule.published_at else None,
            "tags": [{"id": t.id, "tag_name": t.tag_name} for t in (rule.tags or [])],
            "created_at": rule.created_at.isoformat() if rule.created_at else None,
            "updated_at": rule.updated_at.isoformat() if rule.updated_at else None,
        })

    return {
        "code": 200,
        "message": "查詢成功",
        "data": {
            "items": items,
            "total": result["total"],
            "page": result["page"],
            "page_size": result["page_size"],
        },
    }


@router.post("/categories/{category_id}/rules", response_model=dict)
async def create_rule(
    category_id: int,
    data: FaqRuleCreateSchema,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """建立規則"""
    try:
        # 驗證必填欄位
        missing = await faq_service.validate_required_fields(db, category_id, data.content_json)
        if missing:
            raise HTTPException(status_code=400, detail=f"{missing}為必填欄位")

        rule = await faq_service.create_rule(
            db, category_id, data.content_json, data.tag_names, current_user.id
        )
        return {
            "code": 200,
            "message": "規則建立成功",
            "data": {"id": rule.id, "status": rule.status},
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/rules/{rule_id}", response_model=dict)
async def get_rule(
    rule_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """取得單筆規則詳情"""
    rule = await faq_service.get_rule(db, rule_id)
    if not rule:
        raise HTTPException(status_code=404, detail="規則不存在")

    content = rule.content_json
    if isinstance(content, str):
        try:
            content = json.loads(content)
        except (json.JSONDecodeError, TypeError):
            pass

    # 檢查 PMS 唯讀欄位
    from app.services.faq_service import PMS_READONLY_FIELDS
    pms_conn = await faq_service.get_pms_connection(db, rule.category_id)
    pms_readonly = list(PMS_READONLY_FIELDS) if (pms_conn and pms_conn.status == "enabled") else []

    return {
        "code": 200,
        "message": "查詢成功",
        "data": {
            "id": rule.id,
            "category_id": rule.category_id,
            "content_json": content,
            "status": rule.status,
            "created_by": rule.created_by,
            "updated_by": rule.updated_by,
            "published_at": rule.published_at.isoformat() if rule.published_at else None,
            "published_by": rule.published_by,
            "tags": [{"id": t.id, "tag_name": t.tag_name} for t in (rule.tags or [])],
            "created_at": rule.created_at.isoformat() if rule.created_at else None,
            "updated_at": rule.updated_at.isoformat() if rule.updated_at else None,
            "pms_readonly_fields": pms_readonly,
        },
    }


@router.put("/rules/{rule_id}", response_model=dict)
async def update_rule(
    rule_id: int,
    data: FaqRuleUpdateSchema,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """編輯規則"""
    rule = await faq_service.update_rule(
        db, rule_id, data.content_json, data.tag_names, current_user.id
    )
    if not rule:
        raise HTTPException(status_code=404, detail="規則不存在")

    return {
        "code": 200,
        "message": "規則更新成功",
        "data": {"id": rule.id, "status": rule.status},
    }


@router.delete("/rules/{rule_id}", response_model=dict)
async def delete_rule(
    rule_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """刪除規則"""
    deleted = await faq_service.delete_rule(db, rule_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="規則不存在")

    return {"code": 200, "message": "規則刪除成功"}


# === 發佈與版本 ===


@router.post("/rules/{rule_id}/publish", response_model=dict)
async def publish_rule(
    rule_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """發佈規則"""
    rule = await faq_service.publish_rule(db, rule_id, current_user.id)
    if not rule:
        raise HTTPException(status_code=404, detail="規則不存在")

    return {
        "code": 200,
        "message": "規則發佈成功",
        "data": {"id": rule.id, "status": rule.status},
    }



# === 規則狀態切換 ===


@router.patch("/rules/{rule_id}/toggle", response_model=dict)
async def toggle_rule(
    rule_id: int,
    data: FaqRuleToggleSchema,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """切換規則啟用狀態（兩維度：is_enabled + status）"""
    rule = await faq_service.toggle_rule(
        db, rule_id, is_enabled=data.is_enabled, status=data.status
    )
    if not rule:
        raise HTTPException(status_code=404, detail="規則不存在")

    _bump_rule_modified()
    return {
        "code": 200,
        "message": f"規則已{'啟用' if rule.is_enabled else '停用'}",
        "data": {"id": rule.id, "status": rule.status, "is_enabled": rule.is_enabled},
    }


# === 全域發佈 ===


@router.post("/publish", response_model=dict)
async def publish_all(
    token: Optional[str] = Depends(oauth2_scheme),
    db: AsyncSession = Depends(get_db),
):
    """發佈所有 draft 規則（需要 faq.publish 權限）"""
    if not token:
        raise HTTPException(status_code=403, detail="無發佈權限")

    payload = decode_access_token(token)
    if not payload or "sub" not in payload:
        raise HTTPException(status_code=403, detail="無發佈權限")

    user_id = int(payload["sub"])
    result = await db.execute(select(User).where(User.id == user_id))
    current_user = result.scalar_one_or_none()
    if not current_user:
        raise HTTPException(status_code=403, detail="無發佈權限")

    # 檢查 faq_can_publish 權限（admin 或明確授權的用戶）
    if not getattr(current_user, "faq_can_publish", False) and current_user.role.value != "admin":
        raise HTTPException(status_code=403, detail="無發佈權限")

    count = await faq_service.publish_all_draft(db, current_user.id)

    # 廣播「規則已更新」通知至所有聊天室
    if count > 0:
        from app.websocket_manager import manager
        await manager.broadcast("rule_updated", {
            "message": "規則已更新",
            "published_count": count,
        })

    _bump_rule_modified()
    return {
        "code": 200,
        "message": f"已發佈 {count} 筆規則",
        "data": {"published_count": count},
    }


# === 規則最後更新時間（供前端 polling） ===

_last_rule_modified: float = 0.0  # in-memory epoch timestamp

def _bump_rule_modified():
    """更新規則修改時間戳（供 ChatFAB polling 偵測）"""
    global _last_rule_modified
    _last_rule_modified = time.time()

@router.get("/last-modified")
async def get_rules_last_modified():
    """回傳最後一次規則異動的 epoch timestamp"""
    return {"ts": _last_rule_modified}


# === 匯入 / 匯出 ===


EXPORT_HEADERS = ["房型圖片", "房型名稱", "房價", "可入住人數", "剩餘間數", "房型特色", "會員標籤", "訂房 URL"]
EXPORT_DB_KEYS = ["image_url", "房型名稱", "房價", "人數", "間數", "房型特色", None, "url"]


def _build_export_rows(rules) -> list[list[str]]:
    """從 rules 建立匯出用的二維列表（不含標題列）"""
    rows = []
    for rule in rules:
        content = rule.content_json
        if isinstance(content, str):
            try:
                content = json.loads(content)
            except (json.JSONDecodeError, TypeError):
                content = {}
        tag_str = ",".join(t.tag_name for t in (rule.tags or []))
        row = []
        for i, db_key in enumerate(EXPORT_DB_KEYS):
            if db_key is None:
                row.append(tag_str)
            else:
                row.append(str(content.get(db_key, "")))
        rows.append(row)
    return rows


@router.get("/categories/{category_id}/rules/export")
async def export_rules(
    category_id: int,
    format: str = Query("csv", regex="^(csv|xls|xlsx)$"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """匯出分類下所有規則（支援 csv/xls/xlsx）"""
    rules = await faq_service.export_rules(db, category_id)
    rows = _build_export_rows(rules)

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(EXPORT_HEADERS)
        for row in rows:
            writer.writerow(row)
        file_bytes = output.getvalue().encode("utf-8-sig")
        media_type = "text/csv"
        filename = "rules_export.csv"

    elif format == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(EXPORT_HEADERS)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        file_bytes = buf.getvalue()
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "rules_export.xlsx"

    else:  # xls
        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("rules")
        for c, header in enumerate(EXPORT_HEADERS):
            ws.write(0, c, header)
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                ws.write(r + 1, c, val)
        buf = io.BytesIO()
        wb.save(buf)
        file_bytes = buf.getvalue()
        media_type = "application/vnd.ms-excel"
        filename = "rules_export.xls"

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


FIELD_MAPPING = [
    "image_url", "房型名稱", "房價", "人數",
    "間數", "房型特色", "會員標籤", "url",
]
REQUIRED_FIELD_INDICES = {0, 1, 2, 3, 4, 5, 7}  # index 6（會員標籤）允許空值
EXPECTED_FIELD_COUNT = len(FIELD_MAPPING)  # 8


def _clean_excel_value(val) -> str:
    """清理 Excel 讀出的值：None→空字串，浮點整數去小數點"""
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def _validate_and_build_rows(raw_rows: list[list[str]]) -> list[dict[str, str]]:
    """共用校驗：長度檢查、全空列跳過、必填檢查，依固定索引組裝 dict"""
    rows = []
    data_index = 0  # 有效資料的計數（用於錯誤提示）

    for raw in raw_rows:
        # 取前 8 欄（多的忽略）
        values = raw[:EXPECTED_FIELD_COUNT]

        # 全空列 → 靜默跳過
        if not any(v.strip() for v in values):
            continue

        data_index += 1

        # 長度校驗
        if len(values) < EXPECTED_FIELD_COUNT:
            raise ValueError("格式數量不符，請重新匯入")

        # 必填檢查
        for idx in REQUIRED_FIELD_INDICES:
            if not values[idx].strip():
                raise ValueError(
                    f"第 {data_index} 筆資料的「{FIELD_MAPPING[idx]}」為必填欄位"
                )

        # 依固定索引組裝 dict
        row_dict = {FIELD_MAPPING[i]: values[i] for i in range(EXPECTED_FIELD_COUNT)}
        rows.append(row_dict)

    return rows


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    """解析 CSV 檔案"""
    text_content = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text_content))

    # 跳過標題列
    next(reader, None)

    raw_rows = []
    for row in reader:
        raw_rows.append([v.strip() for v in row])

    return _validate_and_build_rows(raw_rows)


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    """解析 .xlsx 檔案"""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # 跳過標題列
    next(rows_iter, None)

    raw_rows = []
    for row_values in rows_iter:
        raw_rows.append([_clean_excel_value(v) for v in row_values])

    wb.close()
    return _validate_and_build_rows(raw_rows)


def _parse_xls(content: bytes) -> list[dict[str, str]]:
    """解析 .xls 檔案"""
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)

    if ws.nrows < 2:
        return []

    # 跳過標題列（row 0），從 row 1 開始
    raw_rows = []
    for r in range(1, ws.nrows):
        raw_rows.append([_clean_excel_value(ws.cell_value(r, c)) for c in range(ws.ncols)])

    return _validate_and_build_rows(raw_rows)


@router.post("/categories/{category_id}/rules/import", response_model=dict)
async def import_rules(
    category_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """匯入規則（以房型名稱為 key，Upsert + 刪除多餘規則）"""
    # 檢查檔案格式
    filename = file.filename or ""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext not in ("csv", "xls", "xlsx"):
        raise HTTPException(
            status_code=400,
            detail="檔案格式不符，僅支援 .csv, .xls, .xlsx",
        )

    content = await file.read()

    try:
        if ext == "csv":
            rows = _parse_csv(content)
        elif ext == "xlsx":
            rows = _parse_xlsx(content)
        else:
            rows = _parse_xls(content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.warning(f"匯入檔案解析失敗: {e}")
        raise HTTPException(status_code=400, detail="檔案解析失敗，請確認檔案內容格式正確")

    if not rows:
        raise HTTPException(status_code=400, detail="檔案中無有效資料")

    count = await faq_service.import_rules(db, category_id, rows, current_user.id)
    _bump_rule_modified()
    return {
        "code": 200,
        "message": f"匯入成功，共 {count} 筆規則",
        "imported_count": count,
    }


# === PMS 串接 ===


async def _test_pms_connection_impl(db: AsyncSession, category_id: int):
    """
    執行 PMS 即時連線測試。
    成功: 回傳 (True, message, room_count)
    失敗: raise HTTPException with 具體錯誤分類
    """
    from app.services.pms_chatbot_client import pms_enabled as pms_configured, query_pms

    if not pms_configured():
        raise HTTPException(
            status_code=400,
            detail="PMS 環境變數未設定（PMS_API_URL / PMS_ACCOUNT / PMS_SECRET / PMS_HOTELCODE）",
        )
    try:
        result = await asyncio.get_event_loop().run_in_executor(
            None, lambda: query_pms("2026-01-01", "2026-01-02")
        )
        room_count = len(result.get("room", []))
        return True, f"連線成功，取得 {room_count} 種房型資料", room_count
    except Exception as e:
        # 記錄錯誤到 DB
        conn = await faq_service.get_pms_connection(db, category_id)
        if conn:
            conn.error_message = str(e)[:500]
            await db.flush()

        msg = str(e)
        if "401" in msg or "Unauthorized" in msg:
            raise HTTPException(status_code=400, detail="連線失敗：API Key 無效（401 Unauthorized）")
        if "403" in msg or "Forbidden" in msg or "whitelist" in msg.lower():
            raise HTTPException(status_code=400, detail="連線失敗：IP 未在白名單，請聯繫 PMS 廠商開通")
        if "timeout" in msg.lower() or "timed out" in msg.lower():
            raise HTTPException(status_code=400, detail="連線失敗：連線逾時，請確認 PMS 端點是否正確")
        raise HTTPException(status_code=400, detail=f"連線失敗：{msg}")


@router.post("/categories/{category_id}/pms-connection/test", response_model=dict)
async def test_pms_connection(
    category_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """測試 PMS 連線（不儲存設定）"""
    success, message, room_count = await _test_pms_connection_impl(db, category_id)
    # 更新 last_synced_at
    conn = await faq_service.get_pms_connection(db, category_id)
    if conn:
        from datetime import datetime
        conn.last_synced_at = datetime.now()
        conn.error_message = None
        await db.flush()
    return {"code": 200, "success": True, "message": message, "room_count": room_count}


@router.post("/categories/{category_id}/pms-connection", response_model=dict)
async def create_pms_connection(
    category_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """建立 PMS 串接設定（先測試連線，成功才儲存）"""
    existing = await faq_service.get_pms_connection(db, category_id)
    if existing:
        raise HTTPException(status_code=400, detail="該大分類已有 PMS 串接設定")

    # 先執行即時連線測試
    await _test_pms_connection_impl(db, category_id)

    conn = await faq_service.create_pms_connection(
        db,
        category_id,
        data.get("api_endpoint", ""),
        data.get("api_key", ""),
        data.get("auth_type", "api_key"),
    )
    return {
        "code": 200,
        "message": "連線測試成功，PMS 串接設定已建立",
        "status": conn.status,
        "auth_type": conn.auth_type,
        "last_synced_at": conn.last_synced_at.isoformat() if conn.last_synced_at else None,
        "snapshot_completed": conn.snapshot_completed,
    }


@router.put("/categories/{category_id}/pms-connection/toggle", response_model=dict)
async def toggle_pms_connection(
    category_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """切換 PMS 串接狀態"""
    conn = await faq_service.toggle_pms_connection(db, category_id, data.get("status", "disabled"))
    if not conn:
        raise HTTPException(status_code=404, detail="PMS 串接設定不存在")

    return {
        "code": 200,
        "message": f"PMS 串接已{'啟用' if conn.status == 'enabled' else '停用'}",
        "status": conn.status,
        "snapshot_completed": conn.snapshot_completed,
    }


@router.get("/categories/{category_id}/pms-connection", response_model=dict)
async def get_pms_connection(
    category_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """查詢 PMS 串接設定"""
    conn = await faq_service.get_pms_connection(db, category_id)
    if not conn:
        raise HTTPException(status_code=404, detail="PMS 串接設定不存在")

    return {
        "status": conn.status,
        "auth_type": conn.auth_type,
        "api_endpoint": conn.api_endpoint,
        "last_synced_at": conn.last_synced_at.isoformat() if conn.last_synced_at else None,
        "snapshot_completed": conn.snapshot_completed,
    }


# === 測試聊天 ===


@router.post("/test-chat", response_model=dict)
async def test_chat(
    data: AiTestChatRequestSchema,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """測試聊天（不計 token、不貼 tag）"""
    result = await ai_chatbot.test_chat(db, data.message)
    return {
        "code": 200,
        "message": "測試完成",
        "data": result,
    }


# === Token 用量 ===


@router.get("/token-usage", response_model=dict)
async def get_token_usage(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """查詢 Token 用量"""
    industry = await faq_service.get_default_industry(db)
    if not industry:
        raise HTTPException(status_code=404, detail="尚未設定產業資料")

    usage = await faq_service.get_token_usage(db, industry.id)
    if not usage:
        return {
            "code": 200,
            "message": "查詢成功",
            "data": {"total_quota": 0, "used_amount": 0, "remaining": 0, "usage_percent": 0},
        }

    remaining = max(0, usage.total_quota - usage.used_amount)
    usage_percent = round(usage.used_amount / usage.total_quota * 100, 1) if usage.total_quota > 0 else 0

    return {
        "code": 200,
        "message": "查詢成功",
        "data": {
            "id": usage.id,
            "industry_id": usage.industry_id,
            "total_quota": usage.total_quota,
            "used_amount": usage.used_amount,
            "remaining": remaining,
            "usage_percent": usage_percent,
        },
    }


@router.put("/token-usage", response_model=dict)
async def update_token_quota(
    data: AiTokenUsageUpdateSchema,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """設定 Token 額度"""
    industry = await faq_service.get_default_industry(db)
    if not industry:
        raise HTTPException(status_code=404, detail="尚未設定產業資料")

    usage = await faq_service.update_token_quota(db, industry.id, data.total_quota)
    return {
        "code": 200,
        "message": "Token 額度設定成功",
        "data": {"total_quota": usage.total_quota},
    }


# === 模組授權 ===


@router.get("/module-auth", response_model=dict)
async def get_module_auth(
    client_id: str = Query(..., description="客戶帳號識別碼"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """查詢模組授權狀態"""
    auth = await faq_service.get_module_auth(db, client_id)
    if not auth:
        return {
            "code": 200,
            "message": "查詢成功",
            "data": {"client_id": client_id, "is_authorized": False},
        }

    return {
        "code": 200,
        "message": "查詢成功",
        "data": {
            "id": auth.id,
            "client_id": auth.client_id,
            "is_authorized": auth.is_authorized,
            "authorized_at": auth.authorized_at.isoformat() if auth.authorized_at else None,
            "authorized_by": auth.authorized_by,
        },
    }


@router.put("/module-auth", response_model=dict)
async def update_module_auth(
    data: FaqModuleAuthUpdateSchema,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """設定模組授權"""
    auth = await faq_service.update_module_auth(
        db, data.client_id, data.is_authorized, current_user.username if hasattr(current_user, 'username') else str(current_user.id)
    )
    return {
        "code": 200,
        "message": f"模組授權已{'開通' if data.is_authorized else '關閉'}",
        "data": {
            "client_id": auth.client_id,
            "is_authorized": auth.is_authorized,
        },
    }
